#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, filedialog
import os
import sys
import json
import platform
import shutil
import subprocess
import urllib.request
import urllib.error
import webbrowser

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("openpyxl is not installed. Run: pip install openpyxl")
    exit(1)

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

COL_WIDTH = 20
BTN_WIDTH_DEL = 4
APP_VERSION = "1.0.3"
REPO_OWNER = "JivkoNushev"
REPO_NAME = "item-price-table-creator"
GITHUB_API = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/releases/latest"
RELEASES_URL = f"https://github.com/{REPO_OWNER}/{REPO_NAME}/releases/latest"

FLOWER_ICON = (
    "iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAAB5UlEQVR4nO2Vz0pCURDG/Yf/xUQD"
    "pU2YoZBP0DYIwmdwFS6McN0DCG2qreADhItWQUSi0MKiNxCyXZDtctNGCPs++YKLVGb3agR34AfH"
    "OXNm5s6ZMzocttjyn+Vpu7JD/ip4BpyIzEKD3z0MnAhaAreiRN0iAruAFwQQtA5eRJ067bnmFdyj"
    "IBGwhKAdMBId6rRHG888ggdBFCQQMA8eDQlwneeebIKWJAGnLLVPX0XHy2AFun0wxHqdcE0d92QT"
    "1TX56OM3gXNgD9TAOWiDK9DQ3Tf55aOuY4yq0NReQ7Ztna3JV26WBFLgALTAs6HUI2PgSSbtdLYl"
    "X6lZq5AGx6D74RBlXQMbXyXAPdoYEujKR3rWWxgLGwmHy+Bm2td/UgWeKZtqRnV/GI5i4BD0fpBA"
    "T7YxnjWbgBuEQFzdXwWDb4IPaKPXENdZt5kEnHqCHDJJdfm08nMqJnUmYHpEa7xG4DgL7tWMWWMz"
    "qvmySoA2WU1Fr6ngSsCloVLVwDE22oUw6oa6hoBl/wtw6AeX4E1BXnUdq6Iu3Ug2tPVbElwJFMA1"
    "6KvLj9RkCRGnTnt92RasTKAITsEZqKi8IT2zsNa8popsaFu0MoEtsAs2+Vuvw6054dHaKdtN2W5Z"
    "loAtttgyT3kHhKXDa84jU2sAAAAASUVORK5CYII="
)


def truncate_text(text, max_chars=COL_WIDTH):
    text = str(text)
    if len(text) <= max_chars:
        return text
    return text[:max_chars - 1] + "\u2026"


def show_msg(parent, title, message, msg_type="info"):
    MsgBox(parent, title, message, msg_type)


class MsgBox(tk.Toplevel):
    def __init__(self, parent, title, message, msg_type="info"):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.after_idle(self.grab_set)

        colors = {"info": "#1565C0", "warning": "#e65100", "error": "#c62828"}
        border_color = colors.get(msg_type, "#1565C0")

        outer = tk.Frame(self, bg=border_color, padx=2, pady=2)
        outer.pack(fill=tk.BOTH, expand=True)

        inner = tk.Frame(outer, bg="white", padx=20, pady=15)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner, text=message, bg="white", font=("Segoe UI", 10),
                 wraplength=350, justify="left", anchor="w").pack(fill=tk.X, pady=(0, 12))

        ok_btn = tk.Button(inner, text="OK", font=("Segoe UI", 10, "bold"),
                           width=10, command=self._close, bg=border_color, fg="white")
        ok_btn.pack()
        ok_btn.focus_set()
        self.bind("<Return>", lambda e: self._close())
        self.bind("<Escape>", lambda e: self._close())

        self.after(10, self._center_on_parent)

    def _center_on_parent(self):
        pw = self.master.winfo_width()
        ph = self.master.winfo_height()
        px = self.master.winfo_x()
        py = self.master.winfo_y()
        w = self.winfo_width()
        h = self.winfo_height()
        self.geometry(f"+{px + (pw - w) // 2}+{py + (ph - h) // 2}")

    def _close(self):
        self.grab_release()
        self.destroy()


class StartupDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Table Generator")
        self.resizable(False, False)
        self.result = None
        self.after_idle(self.grab_set)

        border_color = "#1565C0"
        outer = tk.Frame(self, bg=border_color, padx=2, pady=2)
        outer.pack(fill=tk.BOTH, expand=True)

        inner = tk.Frame(outer, bg="white", padx=24, pady=20)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner, text="Welcome to Table Generator",
                 bg="white", font=("Segoe UI", 12, "bold")).pack(pady=(0, 5))
        tk.Label(inner, text="Choose how to start:",
                 bg="white", font=("Segoe UI", 10)).pack(pady=(0, 15))

        btn_frame = tk.Frame(inner, bg="white")
        btn_frame.pack()

        tk.Button(btn_frame, text="Open File", font=("Segoe UI", 10, "bold"),
                  width=14, bg="#1565C0", fg="white",
                  command=self._open_file).pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(btn_frame, text="Start New", font=("Segoe UI", 10, "bold"),
                  width=14, bg="#4CAF50", fg="white",
                  command=self._start_new).pack(side=tk.LEFT)

        self.protocol("WM_DELETE_WINDOW", self._start_new)
        self.after(10, self._center_on_parent)

    def _center_on_parent(self):
        pw = self.master.winfo_width()
        ph = self.master.winfo_height()
        px = self.master.winfo_x()
        py = self.master.winfo_y()
        w = self.winfo_width()
        h = self.winfo_height()
        self.geometry(f"+{px + (pw - w) // 2}+{py + (ph - h) // 2}")

    def _open_file(self):
        path = filedialog.askopenfilename(
            parent=self, title="Open Entries File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.result = path
            self.grab_release()
            self.destroy()
        else:
            pass

    def _start_new(self):
        self.result = ""
        self.grab_release()
        self.destroy()


class AutocompleteEntry(tk.Frame):
    def __init__(self, parent, suggestions=None, on_search_change=None, **kwargs):
        entry_kwargs = {k: v for k, v in kwargs.items() if k in ("font", "width")}
        frame_kwargs = {k: v for k, v in kwargs.items() if k not in ("font", "width")}
        super().__init__(parent, **frame_kwargs)
        self._suggestions = suggestions if suggestions is not None else []
        self._filtered = []
        self._visible = False
        self._hide_after_id = None
        self.on_search_change = on_search_change

        self.entry = tk.Entry(self, **entry_kwargs)
        self.entry.pack(fill=tk.X, expand=True)
        self.entry.bind("<KeyRelease>", self._on_key_release)
        self.entry.bind("<FocusOut>", self._on_focus_out)
        self.entry.bind("<Down>", self._on_arrow_down)

        root = self.winfo_toplevel()
        self.listbox = tk.Listbox(root, height=6, bd=1, relief=tk.SOLID,
                                  font=entry_kwargs.get("font", ("Segoe UI", 10)))
        self.listbox.bind("<ButtonRelease-1>", self._on_select)
        self.listbox.bind("<Down>", self._on_listbox_down)
        self.listbox.bind("<Up>", self._on_listbox_up)
        self.listbox.bind("<Return>", self._on_select)
        self.listbox.bind("<Escape>", lambda e: self._hide_listbox())
        self.listbox.bind("<Enter>", self._on_listbox_enter)
        self.listbox.bind("<Leave>", self._on_listbox_leave)

    def set_suggestions(self, suggestions):
        self._suggestions = list(suggestions)

    def _update_search(self):
        text = self.entry.get().strip()
        if not text:
            self._filtered = []
            if self.on_search_change:
                self.on_search_change("", [])
            self._hide_listbox()
            return
        lower_text = text.lower()
        self._filtered = [s for s in self._suggestions if lower_text in s.lower()]
        if self.on_search_change:
            self.on_search_change(text, self._filtered)
        if not self._filtered:
            self._hide_listbox()
            return
        self.listbox.delete(0, tk.END)
        for item in self._filtered:
            self.listbox.insert(tk.END, item)
        self._show_listbox()

    def _on_key_release(self, event):
        if event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._update_search()

    def _on_arrow_down(self, event):
        if self._visible and self.listbox.size() > 0:
            self.listbox.focus_set()
            self.listbox.select_set(0)

    def _on_listbox_down(self, event):
        sel = self.listbox.curselection()
        if sel and sel[0] < self.listbox.size() - 1:
            self.listbox.select_clear(0, tk.END)
            self.listbox.select_set(sel[0] + 1)
            self.listbox.see(sel[0] + 1)

    def _on_listbox_up(self, event):
        sel = self.listbox.curselection()
        if sel and sel[0] > 0:
            self.listbox.select_clear(0, tk.END)
            self.listbox.select_set(sel[0] - 1)
            self.listbox.see(sel[0] - 1)

    def _on_select(self, event):
        sel = self.listbox.curselection()
        if sel:
            self.entry.delete(0, tk.END)
            self.entry.insert(0, self.listbox.get(sel[0]))
            self._hide_listbox()
            self.entry.focus_set()
            self._update_search()

    def _on_focus_out(self, event):
        self._hide_after_id = self.after(200, self._hide_listbox)

    def _on_listbox_enter(self, event):
        if self._hide_after_id:
            self.after_cancel(self._hide_after_id)
            self._hide_after_id = None

    def _on_listbox_leave(self, event):
        self._hide_after_id = self.after(200, self._hide_listbox)

    def _show_listbox(self):
        if self._hide_after_id:
            self.after_cancel(self._hide_after_id)
            self._hide_after_id = None
        root = self.winfo_toplevel()
        root_x = root.winfo_rootx()
        root_y = root.winfo_rooty()
        x = self.entry.winfo_rootx() - root_x
        y = self.entry.winfo_rooty() + self.entry.winfo_height() - root_y
        w = self.entry.winfo_width()
        self.listbox.place(x=x, y=y, width=w)
        self.listbox.lift()
        self._visible = True

    def _hide_listbox(self):
        self._hide_after_id = None
        if self._visible:
            self.listbox.place_forget()
            self._visible = False

    def get(self):
        return self.entry.get()

    def delete(self, *args):
        self.entry.delete(*args)

    def insert(self, *args):
        self.entry.insert(*args)


class ColumnDialog(tk.Toplevel):
    def __init__(self, parent, name="", optional=False, title="Add Column"):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.result = None
        self.after(10, self.grab_set)

        border_color = "#1565C0"
        outer = tk.Frame(self, bg=border_color, padx=2, pady=2)
        outer.pack(fill=tk.BOTH, expand=True)

        inner = tk.Frame(outer, bg="white", padx=24, pady=20)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner, text="Column Name:", bg="white",
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        self.name_entry = tk.Entry(inner, font=("Segoe UI", 10), width=30)
        self.name_entry.insert(0, name)
        self.name_entry.pack(pady=(0, 10))
        self.name_entry.focus_set()
        self.name_entry.select_range(0, tk.END)

        self.optional_var = tk.BooleanVar(value=optional)
        tk.Checkbutton(inner, text="Optional", variable=self.optional_var,
                       bg="white", font=("Segoe UI", 10)).pack(anchor="w", pady=(0, 15))

        btn_frame = tk.Frame(inner, bg="white")
        btn_frame.pack()

        tk.Button(btn_frame, text="OK", font=("Segoe UI", 10, "bold"),
                  width=10, bg=border_color, fg="white",
                  command=self._on_ok).pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(btn_frame, text="Cancel", font=("Segoe UI", 10, "bold"),
                  width=10, command=self._on_cancel).pack(side=tk.LEFT)

        self.bind("<Return>", lambda e: self._on_ok())
        self.bind("<Escape>", lambda e: self._on_cancel())
        self.after(10, self._center_on_parent)

    def _center_on_parent(self):
        pw = self.master.winfo_width()
        ph = self.master.winfo_height()
        px = self.master.winfo_x()
        py = self.master.winfo_y()
        w = self.winfo_width()
        h = self.winfo_height()
        self.geometry(f"+{px + (pw - w) // 2}+{py + (ph - h) // 2}")

    def _on_ok(self):
        name = self.name_entry.get().strip()
        if not name:
            show_msg(self, "Warning", "Column name cannot be empty.", "warning")
            return
        self.result = (name, self.optional_var.get())
        self.grab_release()
        self.destroy()

    def _on_cancel(self):
        self.result = None
        self.grab_release()
        self.destroy()


class ColumnBar(tk.Frame):
    def __init__(self, parent, columns, entries, on_change, **kwargs):
        super().__init__(parent, bg="#f0f0f0", **kwargs)
        self.columns = columns
        self.entries = entries
        self.on_change = on_change
        self._build()

    def _build(self):
        for w in self.winfo_children():
            w.destroy()

        for i, col in enumerate(self.columns):
            frame = tk.Frame(self, bg="#D9E1F2", bd=1, relief=tk.RIDGE)
            frame.pack(side=tk.LEFT, padx=2, pady=2)

            tk.Button(frame, text="<", width=2, font=("Segoe UI", 9),
                      command=lambda idx=i: self._move_column(idx, -1)).pack(side=tk.LEFT)

            lbl_text = truncate_text(col["name"], 15)
            if col["optional"]:
                lbl_text += " *"
            name_lbl = tk.Label(frame, text=lbl_text, bg="#D9E1F2",
                                font=("Segoe UI", 9, "bold"), cursor="hand2")
            name_lbl.pack(side=tk.LEFT, padx=2)
            name_lbl.bind("<Button-1>", lambda e, idx=i: self._edit_column(idx))

            tk.Button(frame, text=">", width=2, font=("Segoe UI", 9),
                      command=lambda idx=i: self._move_column(idx, 1)).pack(side=tk.LEFT)

            tk.Button(frame, text="-", width=2, fg="red",
                      font=("Segoe UI", 9, "bold"),
                      command=lambda idx=i: self._delete_column(idx)).pack(side=tk.LEFT)

        tk.Button(self, text="+", width=3, font=("Segoe UI", 10, "bold"),
                  bg="#4CAF50", fg="white",
                  command=self._add_column).pack(side=tk.LEFT, padx=5, pady=2)

    def _add_column(self):
        dialog = ColumnDialog(self.winfo_toplevel())
        self.wait_window(dialog)
        if dialog.result:
            name, optional = dialog.result
            self.columns.append({"name": name, "optional": optional})
            for entry in self.entries:
                entry.append("")
            self.on_change()

    def _delete_column(self, index):
        self.columns.pop(index)
        for entry in self.entries:
            if index < len(entry):
                entry.pop(index)
        self.on_change()

    def _move_column(self, index, direction):
        new_index = index + direction
        if new_index < 0 or new_index >= len(self.columns):
            return
        self.columns[index], self.columns[new_index] = self.columns[new_index], self.columns[index]
        for entry in self.entries:
            if index < len(entry) and new_index < len(entry):
                entry[index], entry[new_index] = entry[new_index], entry[index]
        self.on_change()

    def _edit_column(self, index):
        col = self.columns[index]
        dialog = ColumnDialog(self.winfo_toplevel(), name=col["name"], optional=col["optional"],
                              title="Edit Column")
        self.wait_window(dialog)
        if dialog.result:
            name, optional = dialog.result
            self.columns[index] = {"name": name, "optional": optional}
            self.on_change()


HIGHLIGHT_COLOR = "#C8E6C9"

class EntryRow(tk.Frame):
    def __init__(self, parent, columns, values, delete_cb, save_cb,
                 entries_ref, original_index, is_highlighted=False,
                 base_bg=None, **kwargs):
        super().__init__(parent, bd=1, relief=tk.RIDGE, **kwargs)
        self.columns = columns
        self.values = list(values)
        self.delete_cb = delete_cb
        self.save_cb = save_cb
        self.entries_ref = entries_ref
        self.original_index = original_index
        self._is_highlighted = is_highlighted
        self._base_bg = base_bg or "white"
        self._cells = []
        self._build()

    def _build(self):
        for w in self.winfo_children():
            w.destroy()
        self._cells = []

        tk.Button(self, text="\u2212", width=BTN_WIDTH_DEL, fg="red",
                  font=("Segoe UI", 10, "bold"),
                  command=lambda: self.delete_cb(self)).pack(
                      side=tk.RIGHT, padx=2, pady=2)

        bg = HIGHLIGHT_COLOR if self._is_highlighted else self._base_bg

        for i, val in enumerate(self.values):
            cell = tk.Frame(self, bg=bg)
            cell.pack(side=tk.LEFT, padx=1)
            self._cells.append(cell)
            self._show_label(cell, i, val, bg)

    def _show_label(self, cell, col_idx, value, bg):
        for w in cell.winfo_children():
            w.destroy()
        display = truncate_text(value) if value else "\u2014"
        lbl = tk.Label(cell, text=display, anchor="w", padx=5, pady=4,
                       font=("Segoe UI", 10), width=COL_WIDTH,
                       cursor="hand2", bg=bg)
        lbl.pack()
        lbl.bind("<Button-1>", lambda e, idx=col_idx: self._start_cell_edit(idx))

    def _start_cell_edit(self, col_idx):
        cell = self._cells[col_idx]
        old_value = self.values[col_idx]
        bg = cell.cget("bg")

        for w in cell.winfo_children():
            w.destroy()

        ent = tk.Entry(cell, font=("Segoe UI", 10), width=COL_WIDTH)
        ent.insert(0, old_value)
        ent.pack()
        ent.focus_set()
        ent.select_range(0, tk.END)
        ent.bind("<Return>",
                 lambda e, idx=col_idx, w=ent: self._commit_cell(idx, w))
        ent.bind("<Escape>", lambda e, idx=col_idx: self._cancel_cell(idx))

    def _commit_cell(self, col_idx, entry_widget):
        new_value = entry_widget.get().strip()
        col = self.columns[col_idx]
        if not col["optional"] and not new_value:
            show_msg(self.winfo_toplevel(), "Warning",
                     f'Column "{col["name"]}" cannot be empty.', "warning")
            return
        self.values[col_idx] = new_value
        self.entries_ref[self.original_index] = list(self.values)
        self.save_cb()
        app = self.winfo_toplevel()
        if getattr(app, '_sort_column', -1) >= 0:
            app._render_entries()
            return
        bg = HIGHLIGHT_COLOR if self._is_highlighted else self._base_bg
        self._show_label(self._cells[col_idx], col_idx, new_value, bg)

    def _cancel_cell(self, col_idx):
        bg = HIGHLIGHT_COLOR if self._is_highlighted else self._base_bg
        self._show_label(self._cells[col_idx], col_idx, self.values[col_idx], bg)

    def get_data(self):
        return list(self.values)


def load_entries_from_xlsx(file_path, parent=None):
    columns = []
    entries = []
    if not os.path.exists(file_path):
        return columns, entries
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for cell in row:
                if cell is not None:
                    name = str(cell).strip()
                    if name:
                        columns.append({"name": name, "optional": False})

        if not columns:
            wb.close()
            return columns, entries

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            values = []
            for v in row:
                if v is None:
                    values.append("")
                else:
                    val = str(v).strip()
                    val = val.replace("\u20ac", "").strip()
                    if val in ("\u2014", "-"):
                        val = ""
                    values.append(val)
            while len(values) < len(columns):
                values.append("")
            values = values[:len(columns)]
            entries.append(values)

        for col_idx in range(len(columns)):
            has_empty = any(entry[col_idx] == "" for entry in entries
                            if col_idx < len(entry))
            if has_empty:
                columns[col_idx]["optional"] = True

        wb.close()
    except Exception as e:
        if parent:
            show_msg(parent, "Error", f"Failed to load entries:\n{e}", "error")
    return columns, entries


def save_entries_to_xlsx(columns, entries, file_path, parent=None):
    if not columns:
        return False
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Entries"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                  fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["name"])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, entry in enumerate(entries, 2):
            for col_idx in range(len(columns)):
                val = entry[col_idx] if col_idx < len(entry) else ""
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=val if val else "")
                cell.border = thin_border

        for i in range(len(columns)):
            col_letter = ""
            n = i
            while True:
                col_letter = chr(65 + n % 26) + col_letter
                n = n // 26 - 1
                if n < 0:
                    break
            ws.column_dimensions[col_letter].width = 20

        wb.save(file_path)
        wb.close()
        return True
    except Exception as e:
        if parent:
            show_msg(parent, "Error", f"Failed to save entries:\n{e}", "error")
        return False


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Table Generator")
        self.geometry("750x550")
        self.minsize(650, 450)
        self.configure(bg="#f0f0f0")
        self.current_file = None
        self.columns = []
        self.entries = []
        self.form_entries = []
        self._search_text = ""
        self._search_matches = set()
        self._sort_column = -1
        self._sort_order = 0

        self.iconphoto(True, tk.PhotoImage(data=FLOWER_ICON))
        self._update_title()
        self._build_menu()
        self._build_ui()
        self.after(100, self._show_startup)

    def _build_menu(self):
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open", command=self._open_file, accelerator="Ctrl+O")
        file_menu.add_command(label="Save As", command=self._save_as_file, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.destroy, accelerator="Ctrl+Q")
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Check for Updates...", command=self._check_for_updates)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.config(menu=menubar)

        self.bind_all("<Control-o>", lambda e: self._open_file())
        self.bind_all("<Control-s>", lambda e: self._save_as_file())
        self.bind_all("<Control-q>", lambda e: self.destroy())

    def _build_ui(self):
        style = ttk.Style()
        style.configure("TButton", font=("Segoe UI", 10), padding=4)
        style.configure("TLabel", font=("Segoe UI", 10))

        top_frame = tk.Frame(self, bg="#f0f0f0", pady=5, padx=10)
        top_frame.pack(fill=tk.X)

        colbar_canvas_frame = tk.Frame(top_frame, bg="#f0f0f0")
        colbar_canvas_frame.pack(fill=tk.X, pady=(0, 5))

        self.colbar_canvas = tk.Canvas(colbar_canvas_frame, bg="#f0f0f0",
                                       highlightthickness=0, height=30)
        self.colbar_scrollable = tk.Frame(self.colbar_canvas, bg="#f0f0f0")
        self.colbar_canvas.create_window((0, 0),
            window=self.colbar_scrollable, anchor="nw")
        self.colbar_scrollable.bind("<Configure>",
            lambda e: self.colbar_canvas.configure(
                scrollregion=self.colbar_canvas.bbox("all")))

        self.colbar_canvas.pack(side=tk.TOP, fill=tk.X)

        self.column_bar = ColumnBar(self.colbar_scrollable, self.columns,
                                    self.entries, self._on_columns_change)
        self.column_bar.pack(fill=tk.X)

        form_canvas_frame = tk.Frame(top_frame, bg="#f0f0f0")
        form_canvas_frame.pack(fill=tk.X)

        self.form_canvas = tk.Canvas(form_canvas_frame, bg="#f0f0f0",
                                     highlightthickness=0, height=50)
        self.form_scrollable = tk.Frame(self.form_canvas, bg="#f0f0f0")
        self.form_canvas.create_window((0, 0),
            window=self.form_scrollable, anchor="nw")

        self.form_scrollable.bind("<Configure>",
            lambda e: self.form_canvas.configure(
                scrollregion=self.form_canvas.bbox("all")))

        self.form_canvas.pack(side=tk.TOP, fill=tk.X)

        self.btn_add = tk.Button(top_frame, text="Add Entry",
                                 font=("Segoe UI", 10, "bold"),
                                 bg="#4CAF50", fg="white", padx=12, pady=2,
                                 command=self._add_entry)
        self.btn_add.pack(anchor="w", pady=(5, 0))

        self._rebuild_entry_form()

        sep = ttk.Separator(self, orient="horizontal")
        sep.pack(fill=tk.X, padx=10)

        list_frame = tk.Frame(self, bg="#f0f0f0")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        canvas_frame = tk.Frame(list_frame, bg="#f0f0f0")
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.header_canvas = tk.Canvas(canvas_frame, bg="#D9E1F2",
                                       highlightthickness=0, height=30)
        self.header_scrollable = tk.Frame(self.header_canvas, bg="#D9E1F2")
        self.header_canvas.create_window((0, 0),
            window=self.header_scrollable, anchor="nw")

        self.canvas = tk.Canvas(canvas_frame, bg="white", highlightthickness=0)
        self.vscrollbar = ttk.Scrollbar(canvas_frame, orient="vertical",
                                        command=self._clamped_yview)
        self.hscrollbar = ttk.Scrollbar(canvas_frame, orient="horizontal",
                                        command=self._sync_xview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="white")

        self.scrollable_frame.bind("<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.header_scrollable.bind("<Configure>",
            lambda e: self.header_canvas.configure(
                scrollregion=self.header_canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0),
            window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vscrollbar.set,
                              xscrollcommand=self.hscrollbar.set)
        self.header_canvas.configure(xscrollcommand=self.hscrollbar.set)

        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.hscrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.vscrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.header_canvas.pack(side=tk.TOP, fill=tk.X)
        self.canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self._rebuild_list_header()

        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)

        bottom_frame = tk.Frame(self, bg="#f0f0f0", pady=8, padx=10)
        bottom_frame.pack(fill=tk.X)

        tk.Button(bottom_frame, text="Open", font=("Segoe UI", 10, "bold"),
                  padx=12, pady=2,
                  command=self._open_file).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(bottom_frame, text="Save As", font=("Segoe UI", 10, "bold"),
                  padx=12, pady=2,
                  command=self._save_as_file).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(bottom_frame, text="Finish", font=("Segoe UI", 10, "bold"),
                  bg="#1565C0", fg="white", padx=16, pady=2,
                  command=self._finish).pack(side=tk.RIGHT)

        self._render_entries()

    def _rebuild_entry_form(self):
        for w in self.form_scrollable.winfo_children():
            w.destroy()
        self.form_entries = []
        self.autocomplete = None

        if not self.columns:
            tk.Label(self.form_scrollable,
                     text="Add a column first to start adding entries.",
                     bg="#f0f0f0", font=("Segoe UI", 10, "italic"),
                     fg="gray").pack(anchor="w")
            self.btn_add.config(state=tk.DISABLED)
            return

        self.btn_add.config(state=tk.NORMAL)

        for i, col in enumerate(self.columns):
            frame = tk.Frame(self.form_scrollable, bg="#f0f0f0")
            frame.pack(side=tk.LEFT, padx=(0, 10))

            optional_text = " *" if col["optional"] else ""
            lbl_text = truncate_text(col["name"], COL_WIDTH) + optional_text
            tk.Label(frame, text=lbl_text, bg="#f0f0f0",
                     font=("Segoe UI", 10, "bold")).pack(anchor="w")

            if i == 0:
                entry = AutocompleteEntry(frame, font=("Segoe UI", 10),
                                          width=COL_WIDTH,
                                          on_search_change=self._on_search_change)
                entry.pack(fill=tk.X)
                entry.entry.bind("<Return>",
                                 lambda e: self._focus_next_entry(0))
                self.autocomplete = entry
                self.form_entries.append(entry)
            else:
                entry = tk.Entry(frame, font=("Segoe UI", 10), width=COL_WIDTH)
                entry.pack(fill=tk.X)
                entry.bind("<Return>",
                           lambda e, idx=i: self._focus_next_entry(idx))
                self.form_entries.append(entry)

        if self.form_entries:
            first = self.form_entries[0]
            if isinstance(first, AutocompleteEntry):
                first.entry.focus_set()
            else:
                first.focus_set()

    def _focus_next_entry(self, current_idx):
        next_idx = current_idx + 1
        if next_idx < len(self.form_entries):
            widget = self.form_entries[next_idx]
            if isinstance(widget, AutocompleteEntry):
                widget.entry.focus_set()
            else:
                widget.focus_set()
        else:
            self._add_entry()

    def _rebuild_list_header(self):
        for w in self.header_scrollable.winfo_children():
            w.destroy()

        if not self.columns:
            return

        tk.Label(self.header_scrollable, text="", bg="#D9E1F2",
                 width=BTN_WIDTH_DEL + 2).pack(side=tk.RIGHT)

        for i, col in enumerate(self.columns):
            frame = tk.Frame(self.header_scrollable, bg="#D9E1F2")
            frame.pack(side=tk.LEFT)

            lbl_text = truncate_text(col["name"], COL_WIDTH - 2)
            if col["optional"]:
                lbl_text += " *"

            if self._sort_column == i:
                indicator = "\u2191" if self._sort_order == 1 else "\u2193"
            else:
                indicator = "\u2195"

            sort_lbl = tk.Label(frame, text=indicator, bg="#D9E1F2",
                                font=("Segoe UI", 8), cursor="hand2", padx=1)
            sort_lbl.pack(side=tk.RIGHT)
            sort_lbl.bind("<Button-1>", lambda e, idx=i: self._toggle_sort(idx))

            tk.Label(frame, text=lbl_text, bg="#D9E1F2",
                     font=("Segoe UI", 10, "bold"), anchor="w", padx=3,
                     pady=4, width=COL_WIDTH - 2).pack(side=tk.LEFT)

    def _on_columns_change(self):
        self._sort_column = -1
        self._sort_order = 0
        self.column_bar._build()
        self._rebuild_entry_form()
        self._rebuild_list_header()
        self._refresh_suggestions()
        self._render_entries()
        if self.current_file:
            self._auto_save()

    def _sync_xview(self, *args):
        self.canvas.xview(*args)
        self.header_canvas.xview(*args)
        self.form_canvas.xview(*args)
        self.colbar_canvas.xview(*args)

    def _on_canvas_configure(self, event):
        self.after(10, self._clamp_scroll)

    def _on_mousewheel(self, event):
        x, y = self.winfo_pointerxy()
        widget = self.winfo_containing(x, y)
        if not widget or not self._is_descendant(widget, self.canvas):
            return
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")
        else:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self._clamp_scroll()

    def _clamped_yview(self, *args):
        self.canvas.yview(*args)
        self._clamp_scroll()

    def _clamp_scroll(self):
        content_height = self.scrollable_frame.winfo_reqheight()
        canvas_height = self.canvas.winfo_height()
        if content_height <= canvas_height:
            self.canvas.yview_moveto(0)
        else:
            first, last = self.canvas.yview()
            if first < 0:
                self.canvas.yview_moveto(0)
            elif last > 1:
                self.canvas.yview_moveto(max(0, 1 - (last - first)))

    @staticmethod
    def _is_descendant(widget, ancestor):
        w = widget
        while w:
            if w == ancestor:
                return True
            w = w.master
        return False

    def _refresh_suggestions(self):
        if not self.columns or not self.autocomplete:
            return
        names = list({e[0] for e in self.entries if e})
        names.sort()
        self.autocomplete.set_suggestions(names)

    def _show_startup(self):
        startup = StartupDialog(self)
        self.wait_window(startup)

        if startup.result is None:
            self.destroy()
            return

        self._sort_column = -1
        self._sort_order = 0

        if startup.result:
            self.current_file = startup.result
            cols, entries = load_entries_from_xlsx(self.current_file, self)
            self.columns.clear()
            self.columns.extend(cols)
            self.entries.clear()
            self.entries.extend(entries)
        else:
            self.entries.clear()

        self._update_title()
        self.column_bar._build()
        self._rebuild_entry_form()
        self._rebuild_list_header()
        self._refresh_suggestions()
        self._render_entries()

    def _add_entry(self):
        if not self.columns:
            return

        values = []
        for widget in self.form_entries:
            if isinstance(widget, AutocompleteEntry):
                values.append(widget.get().strip())
            else:
                values.append(widget.get().strip())

        for i, col in enumerate(self.columns):
            if not col["optional"] and not values[i]:
                show_msg(self, "Warning",
                         f'Please enter a value for "{col["name"]}".', "warning")
                return

        if values in self.entries:
            show_msg(self, "Duplicate", "This entry already exists.", "warning")
            return

        self.entries.append(values)

        for widget in self.form_entries:
            if isinstance(widget, AutocompleteEntry):
                widget.delete(0, tk.END)
            else:
                widget.delete(0, tk.END)

        if self.form_entries:
            first = self.form_entries[0]
            if isinstance(first, AutocompleteEntry):
                first.entry.focus_set()
            else:
                first.focus_set()

        self._refresh_suggestions()
        self._render_entries()
        if self.current_file:
            save_entries_to_xlsx(self.columns, self.entries,
                                 self.current_file, self)

    def _finish(self):
        if not self.columns:
            show_msg(self, "Warning", "No columns defined.", "warning")
            return
        if not self.entries:
            show_msg(self, "Warning", "No entries to save.", "warning")
            return
        if not self.current_file:
            path = filedialog.asksaveasfilename(
                title="Save Entries File",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not path:
                return
            self.current_file = path
            self._update_title()
        if save_entries_to_xlsx(self.columns, self.entries,
                                self.current_file, self):
            self.destroy()

    def _auto_save(self):
        if self.current_file and self.columns:
            save_entries_to_xlsx(self.columns, self.entries,
                                 self.current_file, self)

    def _on_search_change(self, text, matches):
        self._search_text = text
        self._search_matches = set(matches) if text else set()
        self._render_entries()

    def _toggle_sort(self, col_idx):
        if self._sort_column != col_idx:
            self._sort_column = col_idx
            self._sort_order = 1
        else:
            self._sort_order = 2 if self._sort_order == 1 else 0
            if self._sort_order == 0:
                self._sort_column = -1
        self._rebuild_list_header()
        self._render_entries()

    def _get_sorted_entries(self):
        indexed = list(enumerate(self.entries))
        if self._sort_column >= 0 and self._sort_order != 0:
            reverse = self._sort_order == 2
            indexed.sort(key=lambda x: str(x[1][self._sort_column]).lower()
                         if self._sort_column < len(x[1]) else "", reverse=reverse)
        return indexed

    def _render_entries(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        if not self.entries:
            text = "No entries yet." if self.columns else "No columns defined."
            tk.Label(self.scrollable_frame, text=text, bg="white",
                     font=("Segoe UI", 10, "italic"), fg="gray",
                     pady=20).pack()
            self.after(10, self._clamp_scroll)
            return

        for display_idx, (orig_idx, entry_values) in enumerate(self._get_sorted_entries()):
            is_highlighted = bool(self._search_text) and entry_values[0] in self._search_matches
            base_bg = "white" if display_idx % 2 == 0 else "#f9f9f9"
            row_bg = HIGHLIGHT_COLOR if is_highlighted else base_bg
            row = EntryRow(self.scrollable_frame, self.columns, entry_values,
                           delete_cb=self._delete_entry,
                           save_cb=self._auto_save,
                           entries_ref=self.entries,
                           original_index=orig_idx,
                           is_highlighted=is_highlighted,
                           base_bg=base_bg,
                           bg=row_bg)
            row.pack(fill=tk.X)
        self.after(10, self._clamp_scroll)

    def _delete_entry(self, row_widget):
        data = row_widget.get_data()
        self.entries[:] = [e for e in self.entries if e != data]
        self._refresh_suggestions()
        self._render_entries()
        if self.current_file:
            save_entries_to_xlsx(self.columns, self.entries,
                                 self.current_file, self)

    def _update_title(self):
        base = f"Table Generator v{APP_VERSION}"
        if self.current_file:
            self.title(f"{base} \u2014 {os.path.basename(self.current_file)}")
        else:
            self.title(f"{base} \u2014 New")

    def _parse_version(self, v_str):
        return tuple(int(x) for x in v_str.lstrip("vV").split("."))

    def _check_for_updates(self):
        try:
            req = urllib.request.Request(GITHUB_API, headers={"Accept": "application/vnd.github.v3+json", "User-Agent": "TableGenerator"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            if e.code == 404:
                show_msg(self, "Update Check", "No releases found on GitHub yet.\nPush to master to trigger a release build.", "info")
            else:
                show_msg(self, "Update Check", f"Update server returned an error ({e.code}).", "warning")
            return
        except Exception:
            show_msg(self, "Update Check", "Could not check for updates.\nCheck your internet connection and try again.", "warning")
            return

        latest_tag = data.get("tag_name", "")
        if not latest_tag:
            show_msg(self, "Update Check", "No release information found.", "info")
            return

        current = self._parse_version(APP_VERSION)
        latest = self._parse_version(latest_tag)

        if latest <= current:
            show_msg(self, "Update Check", f"You're up to date (v{APP_VERSION}).", "info")
            return

        body = data.get("body", "")
        self._show_update_dialog(latest_tag, body)

    def _show_update_dialog(self, latest_tag, body):
        win = tk.Toplevel(self)
        win.title("Update Available")
        win.resizable(False, False)
        win.after_idle(win.grab_set)

        outer = tk.Frame(win, bg="#4CAF50", padx=2, pady=2)
        outer.pack(fill=tk.BOTH, expand=True)
        inner = tk.Frame(outer, bg="white", padx=20, pady=15)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner, text="Update Available", font=("Segoe UI", 12, "bold"),
                 bg="white", fg="#4CAF50").pack(anchor="w", pady=(0, 10))

        lines = [
            f"Current version:  v{APP_VERSION}",
            f"New version:      {latest_tag}",
        ]
        for line in lines:
            tk.Label(inner, text=line, bg="white", font=("Segoe UI", 10),
                     anchor="w", justify="left").pack(anchor="w")

        if body:
            tk.Frame(inner, bg="#e0e0e0", height=1).pack(fill=tk.X, pady=8)
            tk.Label(inner, text="What's new:", bg="white",
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(anchor="w")
            tk.Label(inner, text=body[:500], bg="white", font=("Segoe UI", 9),
                     wraplength=380, justify="left", anchor="w").pack(anchor="w", pady=(2, 10))

        btn_frame = tk.Frame(inner, bg="white")
        btn_frame.pack(fill=tk.X, pady=(8, 0))

        tk.Button(btn_frame, text="Download & Install", font=("Segoe UI", 10, "bold"),
                  bg="#4CAF50", fg="white", padx=12,
                  command=lambda: self._download_and_install(win)).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_frame, text="Cancel", font=("Segoe UI", 10),
                  command=win.destroy).pack(side=tk.RIGHT)
        win.bind("<Escape>", lambda e: win.destroy())

        win.after(10, lambda: self._center_window(win))

    def _center_window(self, win):
        pw, ph = self.winfo_width(), self.winfo_height()
        px, py = self.winfo_x(), self.winfo_y()
        w, h = win.winfo_width(), win.winfo_height()
        win.geometry(f"+{px + (pw - w) // 2}+{py + (ph - h) // 2}")

    def _download_and_install(self, dialog):
        dialog.destroy()
        system = platform.system()

        if system == "Windows":
            asset_name = "TableGenerator.exe"
        else:
            asset_name = "main.py"

        try:
            req = urllib.request.Request(GITHUB_API, headers={"Accept": "application/vnd.github.v3+json", "User-Agent": "TableGenerator"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
        except Exception:
            show_msg(self, "Error", "Failed to fetch update info.", "error")
            return

        asset_url = None
        for asset in data.get("assets", []):
            if asset["name"] == asset_name:
                asset_url = asset["browser_download_url"]
                break

        if not asset_url:
            show_msg(self, "Error", f"No {asset_name} found in the latest release.", "error")
            return

        progress = tk.Toplevel(self)
        progress.title("Downloading")
        progress.resizable(False, False)
        progress.after_idle(progress.grab_set)
        tk.Label(progress, text="Downloading update...", font=("Segoe UI", 10),
                 padx=30, pady=20).pack()
        progress.update()

        temp_dir = os.path.join(os.path.expanduser("~"), ".table_generator_update")
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, asset_name)

        try:
            urllib.request.urlretrieve(asset_url, temp_path)
        except Exception:
            progress.destroy()
            show_msg(self, "Error", "Download failed. Please try again.", "error")
            return

        progress.destroy()

        restart = tk.Toplevel(self)
        restart.title("Update Ready")
        restart.resizable(False, False)
        restart.after_idle(restart.grab_set)

        outer = tk.Frame(restart, bg="#1565C0", padx=2, pady=2)
        outer.pack(fill=tk.BOTH, expand=True)
        inner = tk.Frame(outer, bg="white", padx=20, pady=15)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner, text="Update downloaded successfully.", font=("Segoe UI", 10),
                 bg="white").pack(pady=(0, 12))

        btn_frame = tk.Frame(inner, bg="white")
        btn_frame.pack()
        tk.Button(btn_frame, text="Restart Now", font=("Segoe UI", 10, "bold"),
                  bg="#1565C0", fg="white", padx=12,
                  command=lambda: self._install_update(temp_path, restart)).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn_frame, text="Later", font=("Segoe UI", 10),
                  command=restart.destroy).pack(side=tk.LEFT)

        restart.bind("<Escape>", lambda e: restart.destroy())

    def _get_app_path(self):
        if getattr(sys, "frozen", False):
            return sys.executable
        return os.path.abspath(sys.argv[0])

    def _install_update(self, temp_path, dialog):
        dialog.destroy()
        app_path = self._get_app_path()
        system = platform.system()

        if system == "Windows":
            update_bat = os.path.join(os.path.dirname(temp_path), "update.bat")
            with open(update_bat, "w") as f:
                f.write(f'@echo off\n')
                f.write(f'timeout /t 2 /nobreak >nul\n')
                f.write(f'copy /y "{temp_path}" "{app_path}" >nul 2>&1\n')
                f.write(f'if errorlevel 1 (\n')
                f.write(f'  echo Update failed. Run manually: copy "{temp_path}" "{app_path}"\n')
                f.write(f'  pause\n')
                f.write(f'  exit /b 1\n')
                f.write(f')\n')
                f.write(f'start "" "{app_path}"\n')
                f.write(f'del "%~f0"\n')
            subprocess.Popen(["cmd.exe", "/c", "start", "/min", update_bat],
                             shell=True, close_fds=True)
            self.destroy()
        else:
            try:
                shutil.move(temp_path, app_path)
            except Exception:
                show_msg(self, "Error", "Could not write the update.\nMake sure you have write permissions.", "error")
                return
            os.execv(sys.executable, [sys.executable, app_path])

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Open Entries File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self.current_file = path
        cols, entries = load_entries_from_xlsx(self.current_file, self)
        self.columns.clear()
        self.columns.extend(cols)
        self.entries.clear()
        self.entries.extend(entries)
        self._sort_column = -1
        self._sort_order = 0
        self._update_title()
        self.column_bar._build()
        self._rebuild_entry_form()
        self._rebuild_list_header()
        self._refresh_suggestions()
        self._render_entries()

    def _save_as_file(self):
        path = filedialog.asksaveasfilename(
            title="Save Entries File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self.current_file = path
        self._update_title()
        save_entries_to_xlsx(self.columns, self.entries, self.current_file, self)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
